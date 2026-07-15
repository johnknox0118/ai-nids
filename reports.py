import os, csv, datetime
from config import REPORTS_DIR
import database as db


def _ensure():
    os.makedirs(REPORTS_DIR, exist_ok=True)


def _save_record(name, path, total):
    try:
        conn = db.get_db()
        conn.execute("INSERT INTO reports (report_name,total_attacks,file_path) VALUES (?,?,?)",
                     (name, total, path))
        conn.commit()
        conn.close()
    except Exception:
        pass


def generate_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    _ensure()
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    fn = f"NIDS_Report_{ts}.pdf"
    fp = os.path.join(REPORTS_DIR, fn)
    doc = SimpleDocTemplate(fp, pagesize=A4,
                            rightMargin=0.5*inch, leftMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    H1 = ParagraphStyle('H1', fontSize=20, fontName='Helvetica-Bold',
                         textColor=colors.HexColor('#1a73e8'), spaceAfter=6)
    H2 = ParagraphStyle('H2', fontSize=13, fontName='Helvetica-Bold',
                         textColor=colors.HexColor('#333333'), spaceAfter=4)

    def tbl(data, hc):
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor(hc)),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.HexColor('#f8f9fa'),colors.white]),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#dee2e6')),
            ('FONTSIZE',(0,0),(-1,-1),8), ('PADDING',(0,0),(-1,-1),5),
        ]))
        return t

    stats = db.get_stats()
    elems = [
        Paragraph("AI-NIDS Security Report", H1),
        Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor('#1a73e8')),
        Spacer(1, 10),
        Paragraph("Executive Summary", H2),
        tbl([['Metric','Value'],
             ['Total Packets', str(stats['total_packets'])],
             ['Total Threats', str(stats['total_threats'])],
             ["Today's Alerts", str(stats['today_alerts'])],
             ['Blocked IPs', str(stats['blocked_ips'])]], '#1a73e8'),
        Spacer(1, 12),
    ]
    attacks = db.get_attack_frequency()
    if attacks:
        elems += [Paragraph("Attack Frequency", H2),
                  tbl([['Attack','Count']]+[[a['attack_name'],str(a['cnt'])] for a in attacks],'#dc3545'),
                  Spacer(1, 12)]
    threats = db.get_recent_threats(20)
    if threats:
        rows = [['Attack','Severity','Source IP','Time','Recommendation']]
        for t in threats:
            rows.append([t['attack_name'], t['severity'], t['source_ip'] or '-',
                         (t['time'] or '')[:16], (t['recommendation'] or '')[:45]])
        elems += [Paragraph("Recent Threats", H2), tbl(rows, '#6f42c1')]

    doc.build(elems)
    _save_record(fn, fp, stats['total_threats'])
    return fp, fn


def generate_csv():
    _ensure()
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    fn = f"NIDS_Threats_{ts}.csv"
    fp = os.path.join(REPORTS_DIR, fn)
    threats = db.get_recent_threats(1000)
    fields = ['id','attack_name','severity','source_ip','destination_ip',
              'protocol','time','recommendation','status']
    with open(fp, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        w.writerows(threats)
    _save_record(fn, fp, len(threats))
    return fp, fn


def generate_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        _ensure()
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        fn = f"NIDS_Report_{ts}.xlsx"
        fp = os.path.join(REPORTS_DIR, fn)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Threats"
        headers = ['ID','Attack','Severity','Source IP','Dest IP',
                   'Protocol','Time','Recommendation','Status']
        fill = PatternFill("solid", fgColor="1a73e8")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        for row, t in enumerate(db.get_recent_threats(500), 2):
            for col, key in enumerate(['id','attack_name','severity','source_ip',
                                       'destination_ip','protocol','time',
                                       'recommendation','status'], 1):
                ws.cell(row=row, column=col, value=t.get(key, ''))
        ws2 = wb.create_sheet("Summary")
        ws2.append(['Metric','Value'])
        for k, v in db.get_stats().items():
            ws2.append([k.replace('_',' ').title(), v])
        wb.save(fp)
        _save_record(fn, fp, db.get_stats()['total_threats'])
        return fp, fn
    except Exception:
        return generate_csv()
